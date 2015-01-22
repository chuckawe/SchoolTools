# For reading XLS files
#from xlrd import open_workbook
# For reading XLSX files
from openpyxl.reader.excel import load_workbook
# For reading Goodle Docs
import gdata.docs
import gdata.docs.service
import gdata.spreadsheet.service

# Imports from this package
from Exam import *
from Student import *
from Course import *
from Section import *
from Behavior import *

# Other python imports
import getpass
from string import atoi
import os
from os.path import expanduser
import datetime,time

class Reader(object):
    """ """
    def __init__(self,excelFile):
        self.excelFile = excelFile
        #self.b = open_workbook(self.excelFile)
        self.b = load_workbook(filename = self.excelFile)
        
        # Get first worksheet (assume this is what we want)
        #self.s = self.b.sheet_by_index(0)
        self.s = self.b.worksheets[0]
        self.current_sheet = 0
    
    def next_sheet(self):
        if (self.current_sheet) == len(self.b.worksheets): return 0
        
        self.s = self.b.worksheets[self.current_sheet]
        print "Ready to open sheet:",self.s.title
        self.current_sheet += 1

        return 1

    def set_sheet(self,title):
        self.current_sheet=0
        for sheet in self.b.worksheets:
            if sheet.title == title:
                self.s = sheet
                print "Ready to open sheet:",self.s.title
                return 0
            self.current_sheet += 1
        
        return 1


    def find_id_cell(self):
        """returns row and column indexes of ID column header, fatal error if not found"""
        for row in range(0,self.s.get_highest_row()):
            for col in range(0,self.s.get_highest_column()):
                val = self.s.cell(row=row+1,column=col+1).value
                if self.s.cell(row=row+1,column=col+1).data_type == 's' and val is not None:
                    if not set(['id','stuid','studentid']).isdisjoint(set(val.lower().split(' '))):
                        #print 'Found row containing student IDs stored in: ', val
                        return row+1,col+1
        
        print 'Problem parsing Excel file; cannot find column storing student ID'
        import sys
        sys.exit(1)


    def find_id_column(self):
        """returns index of header column."""
        row,col = self.find_id_cell()
        return col

    def find_header_row(self):
        """returns index of header row."""
        row,col = self.find_id_cell()
        return row
            
    def find_column(self, keys):
        if self.header_row is None:
            self.header_row = self.find_header_row()
        for col in range(0,self.s.get_highest_column()):
            val = self.s.cell(row=self.header_row,column=col+1).value
            if self.s.cell(row=self.header_row,column=col+1).data_type == 's' and val is not None:
                if not set(keys).isdisjoint(set(val.lower().split(' '))):
                    return col+1
    


    def get_ids(self):
        """Returns list of student ids in file."""
        ids = []
        head_row,id_col = self.find_id_cell()
        
        for row in range(head_row+1,self.s.get_highest_row()):
            val = self.s.cell(row=row,column=id_col).value
            if self.s.cell(row=row,column=id_col).data_type == 's':
                # remove hyphens
                val = val.replace('-','')
                ids.append(atoi(val))
            else:
                ids.append(int(val))
        
        return ids

class OldStudentDataReader(Reader):
    def __init__(self,excelFile):
        super(StudentDataReader, self).__init__(excelFile)

        self.header_row,self.id_column = self.find_id_cell()
        self.last_column  = self.find_column(['lastname'])
        self.first_column  = self.find_column(['firstname'])
        self.gradYear_column  = self.find_column(['gradclass'])
        self.street_column  = self.find_column(['address1'])
        self.city_column  = self.find_column(['city1'])
        self.state_column  = self.find_column(['state1'])
        self.zipcode_column  = self.find_column(['zip1'])
        self.dob_column  = self.find_column(['birthdate'])
        self.iep_column  = self.find_column(['iep'])
        self.p504_column  = self.find_column(['504plan'])
        self.ell_column  = self.find_column(['ell'])

    def get_data(self,students,get_all=True, year=-1): 
        """Returns list of students with student data information"""
        
        print 'From year:', (str(year)+'-'+str(year+1)) 
        print 'From title:',self.s.title
        # Check if year is requested year
        if year != -1 and (str(year)+'-'+str(year+1)) != self.s.title: 
            return

        for row in range(self.header_row+1,self.s.get_highest_row()+1):
            rowID = 0
            val1 = self.s.cell(row=row,column=self.id_column).value
            if self.s.cell(row=row,column=self.id_column).data_type == 's':
                # remove hyphens
                val1 = val1.replace('-','')
                rowID = atoi(val1)
            else:
                rowID = int(val1)
            
            student = find_by_id(students,rowID)
            if student is None:
                if get_all:
                    student = Student()
                    students.append(student)
                else:
                    continue
            
            # Get ID and year
            student.stuID    = rowID
            student.gradYear = int(self.s.cell(row=row,column=self.gradYear_column).value)
            # Get Name
            student.last   = str(self.s.cell(row=row,column=self.last_column).value)
            student.first  = str(self.s.cell(row=row,column=self.first_column).value)
            # Get Address
            student.street  = str(self.s.cell(row=row,column=self.street_column).value)
            student.city    = str(self.s.cell(row=row,column=self.city_column).value)
            student.state   = str(self.s.cell(row=row,column=self.state_column).value)
            student.zipcode = str(self.s.cell(row=row,column=self.zipcode_column).value)
            # DOB
            if str(self.s.cell(row=row,column=self.dob_column).value) == "None":
                student.dob = ""
            else:
                time_tuple  = time.strptime(str(self.s.cell(row=row,column=self.dob_column).value),"%Y-%m-%d %H:%M:%S")
                student.dob = str(time_tuple.tm_mon)+"/"+str(time_tuple.tm_mday)+"/"+str(time_tuple.tm_year)
			# Get IEP/504/ELL
            student.HasIEP   = (str(self.s.cell(row=row,column=self.iep_column).value) == 'IEP')
            student.Has504   = (str(self.s.cell(row=row,column=self.p504_column).value) == '504 Plan')
            student.IsELL   = (str(self.s.cell(row=row,column=self.ell_column).value) == 'ELL')

class StudentDataReader(Reader):
    def __init__(self,excelFile):
        super(StudentDataReader, self).__init__(excelFile)

        self.header_row,self.id_column = self.find_id_cell()
        self.last_column  = self.find_column(['lastname'])
        self.first_column  = self.find_column(['firstname'])
        self.grade_column  = self.find_column(['grade'])
        self.dob_column  = self.find_column(['dob'])
        self.iep_column  = self.find_column(['iep'])
        self.p504_column  = self.find_column(['504plan'])
        self.ell_column  = self.find_column(['ell'])

    def get_data(self,students,get_all=True, year=-1): 
        """Returns list of students with student data information"""
        
        print 'From year:', (str(year)+'-'+str(year+1)) 
        print 'From title:',self.s.title
        # Check if year is requested year
        if year != -1 and (str(year)+'-'+str(year+1)) != self.s.title: 
            return

        for row in range(self.header_row+1,self.s.get_highest_row()+1):
            rowID = 0
            val1 = self.s.cell(row=row,column=self.id_column).value
            if self.s.cell(row=row,column=self.id_column).data_type == 's':
                # remove hyphens
                val1 = val1.replace('-','')
                rowID = atoi(val1)
            else:
                rowID = int(val1)
            
            student = find_by_id(students,rowID)
            if student is None:
                if get_all:
                    student = Student()
                    students.append(student)
                else:
                    continue
            
            # Get ID and year
            student.stuID    = rowID
            student.set_hs_grade(int(self.s.cell(row=row,column=self.grade_column).value))
            # Get Name
            student.last   = str(self.s.cell(row=row,column=self.last_column).value)
            student.first  = str(self.s.cell(row=row,column=self.first_column).value)
            # DOB
            if str(self.s.cell(row=row,column=self.dob_column).value) == "None":
                student.dob = ""
            else:
                time_tuple  = time.strptime(str(self.s.cell(row=row,column=self.dob_column).value),"%Y-%m-%d %H:%M:%S")
                student.dob = str(time_tuple.tm_mon)+"/"+str(time_tuple.tm_mday)+"/"+str(time_tuple.tm_year)
	    # Get IEP/504/ELL
            #student.HasIEP   = (str(self.s.cell(row=row,column=self.iep_column).value) == 'IEP')
            #student.Has504   = (str(self.s.cell(row=row,column=self.p504_column).value) == '504 Plan')
            #student.IsELL   = (str(self.s.cell(row=row,column=self.ell_column).value) == 'ELL')



class CourseReader(Reader):
    def __init__(self,excelFile):
        super(CourseReader, self).__init__(excelFile)

        # Find the columns containing the needed information
        self.header_row,self.id_column = self.find_id_cell()
        self.title_column     = self.find_column(['subject'])
        self.grade_column     = self.find_column(['grade'])
        self.year_column      = self.find_column(['year'])
        self.credits_column   = self.find_column(['credits'])
        self.level_column     = self.find_column(['level'])
        self.summer_column    = self.find_column(['passsummer'])
    
    def get_courses(self,students):
        for row in range(self.header_row+1,self.s.get_highest_row()+1):
            rowID = 0 
            val1 = self.s.cell(row=row,column=self.id_column).value
            if self.s.cell(row=row,column=self.id_column).data_type == 's':
                # remove hyphens
                val1 = val1.replace('-','')
                rowID = atoi(val1)
            else:
                rowID = int(val1)
            
            student = find_by_id(students,rowID)
            if student is not None:
                title = str(self.s.cell(row=row,column=self.title_column).value)
                if 'Enrolled' in title: continue
                if 'Creative Reading' in title: continue
                course = Course()
                # Get Grade
                course.title = title 
                course.numAvg  = float(self.s.cell(row=row,column=self.grade_column).value)
                course.nCredits = float(self.s.cell(row=row,column=self.credits_column).value)
                course.year = int(self.s.cell(row=row,column=self.year_column).value)
                course.level = int(self.s.cell(row=row,column=self.level_column).value)
                course.PassSummer = False
                if 'yes' in str(self.s.cell(row=row,column=self.summer_column).value).lower():
                    course.PassSummer = True
                student.add_course(course)

class CourseCurrentReader(Reader):
    def __init__(self,excelFile):
        super(CourseCurrentReader, self).__init__(excelFile)

        # Find the columns containing the needed information
        self.header_row,self.id_column = self.find_id_cell()
        self.title_column     = self.find_column(['subject'])
        self.grade_column     = self.find_column(['percent'])
        #self.year_column      = self.find_column(['year'])
        #self.credits_column   = self.find_column(['credits'])
        #self.level_column     = self.find_column(['level'])
        #self.summer_column    = self.find_column(['passsummer'])
        self.term_column     = self.find_column(['term'])
    
    def get_courses(self,students,final_term):
        now = datetime.datetime.now()
        year = now.year
        if now.month < 8:
            year = year - 1
        print 'Reading current courses and setting year as',year

        # For 2012-2013 here are the course credit aslottments
        # This is used to determine the number of credits for 
        # current-year courses on-the-fly
        #courses_3credit = {
        #'Biology','Integrated Algebra','Korean I',
        #'Chemistry','Geometry','Global History II: Enlightenment to Present Day',
        #'Physics','Algebra II','U.S. History','Korean III','Algebra II/Trigonometry',
        #'AP Biology','Intro to Calculus','Pre-Calculus','Senior Literature','AP Eng Lang \& Comp',

        #'Literature I','Writing I','Global History I: Mesopotamia to Renaissance',
        #'Literature II','Writing II','Korean II','American Literature',
        #'Literature III','Writing III','AP Physics',
        #'Korean Literature & Culture','Korean IV - Regents Prep',
        #'Korean Lit. \& Culture'
        #}
        #courses_2credit = [
        #'Applied Civics II: Seminar in American Democracy',
        #'Sem in Amer Democ'
        #]
        #courses_1credit = [
        #'Physics Topics','Scientific Literacy',
        #'Topics in Physics','Economics'
        #]
        #courses_5credit = [
        #'Theatre for Social Change with The Laramie Project','Physical Education','Advanced Theatre'
        #]

        for row in range(self.header_row+1,self.s.get_highest_row()+1):
            term = str(self.s.cell(row=row,column=self.term_column).value)
            # Check that this row corresponds to requested term
            if term not in ['F1','F2','F3']:
            #if term not in ['T1','T2','T3']:
                continue
                
            rowID = 0 
            val1 = self.s.cell(row=row,column=self.id_column).value
            if self.s.cell(row=row,column=self.id_column).data_type == 's':
                # remove hyphens
                val1 = val1.replace('-','')
                rowID = atoi(val1)
            else:
                rowID = int(val1)
            
            student = find_by_id(students,rowID)
            #if student is None:
                #print 'Warning: Missing student with ID =',rowID  
            #else:
            if student is not None:
                title = str(self.s.cell(row=row,column=self.title_column).value)
                #title = title.replace('&','\&')  # Protect against naughty character
                #title = title.replace('Korean Literature','Korean Lit.')  # Shorten Kor Lit Title
                #title = title.replace('Korean IV - Regents Prep','Korean IV')  # Shorten Kor IV Title
                #title = title.replace('Applied Civics II: ','')  # Shorten Civics Seminar
                #title = title.replace('Seminar in American Democracy','Sem in Amer Democ')  # Shorten Civics Seminar
                #title = title.replace('AP English Language and Composition','AP Eng Lang \& Comp')  # Shorten AP English

                if self.s.cell(row=row,column=self.grade_column).value == None:
                    continue
                course = Course()
                # Check if Course already in for different term
                for c in student.get_courses(year):
                    if c.title == title:
                        course = c
                    
                # Fill course information
                if len(course.term_grades) == 0:
                    course.title = title 
                    course.current = True
                    course.year = year
                    if 'Hon' in title:
                        course.level = 2
                    elif 'AP' in title:
                        course.level = 3
                    else:
                        course.level = 1
                
                # Fill grade
                course.term_grades[term] = float(self.s.cell(row=row,column=self.grade_column).value)
                if term == final_term:
                    course.numAvg  = float(self.s.cell(row=row,column=self.grade_column).value)

                #stripTitle = title.replace(' (Hon)','')
                #stripTitle = stripTitle.replace(' Accelerated','')
                #if stripTitle in courses_3credit:
                #    course.nCredits = 3
                #elif stripTitle in courses_2credit:
                #    course.nCredits = 2
                #elif stripTitle in courses_1credit:
                #    course.nCredits = 1
                #elif stripTitle in courses_5credit:
                #    course.nCredits = 0.5
                #else:
                #    course.nCredits = 0
                    
                
                if len(course.term_grades) == 1:
                    student.add_course(course)

class BehaviorReader(Reader):
    def __init__(self,excelFile):
        super(BehaviorReader, self).__init__(excelFile)
        
        # Find the columns containing the needed information
        self.header_row,self.id_column = self.find_id_cell()
        self.behavior_column  = self.find_column(['behavior']) 
        self.category_column = self.find_column(['category'])
        self.date_column   = self.find_column(['date'])

    def get_behaviors(self,students):
        """Return list students with behaviors loaded"""
        pass
        #for row in range(self.header_row+1,self.s.get_highest_row()+1):

    def get_data(self):
        '''Returns simple container with behaviors'''
        
        behaviors = []
        
        for row in range(self.header_row+1,self.s.get_highest_row()+1):
            stuid = int(self.s.cell(row=row,column=self.id_column).value)
            date = self.s.cell(row=row,column=self.date_column).value
            behav = self.s.cell(row=row,column=self.behavior_column).value
            cat = self.s.cell(row=row,column=self.category_column).value
            # Determine Behavior Type
            behaviorType = ''
            if cat == 'Demeritable Behaviors':
                behaviorType = 'Demerit'
            elif behav == 'Sent out':
                behaviorType = 'SendOut'
            elif cat == 'Auto-Detention' and behav == 'Uniform':
                behaviorType = 'Uniform'
            elif cat == 'Auto-Detention' and behav == 'Automatic Detention - Please Note':
                behaviorType = 'AutoDetention'
            else:
                continue

            behaviors.append(Behavior(behaviorType=behaviorType,date=date,student=stuid))
                
        return behaviors 
                
                

            
        


class RegentsReader(Reader):
    
    def __init__(self,excelFile):
        super(RegentsReader, self).__init__(excelFile)
        
        # Regents Subject Codes
        # These are from ATS.  For some courses, depending on the year
        # there are multiple codes for a given course.  Also, a fifth letter
        # may be appended to the 4 letter code.  This indicates the term
        # in which the exam was taken.
        self.code_dict = {
            'SXRK':'Living Environment',
            'SXZK':'Living Environment',
            'SXRX':'Chemistry',
            'SXRC':'Chemistry',
            'SXRP':'Physics',
            'MXRA':'Mathematics A',
            'MXZC':'Mathematics',
            'MXRE':'Integrated Algebra',
            'MXZE':'Integrated Algebra',
            'MXRC':'Algebra I (Common Core)',
            'MXRG':'Geometry',
            'MXRT':'Algebra II/Trigonometry',
            'HXRU':'US History',
            'HXRA':'US History',
            'HXRG':'Global History',
            'HXR$':'Global History',
            'SXRU':'Earth Science',
            'EXRL':'English',
            'FXRS':'Spanish',
            'FXPS':'Spanish',
            'FXRF':'French',
            'FXRT':'Italian',
            'KLOT':'Korean'
        }
        

        self.header_row,self.id_column = self.find_id_cell()
        self.score_column  = self.find_column(['mark','finalscore','final']) 
        self.course_column = self.find_column(['code'])
        self.year_column   = self.find_column(['year'])
        self.month_column  = self.find_column(['term'])


    def get_exams(self,students):
        """Return list of Exam objects for a particular student"""
        
        for row in range(self.header_row+1,self.s.get_highest_row()+1):
            rowID = 0 
            val1 = self.s.cell(row=row,column=self.id_column).value
            if self.s.cell(row=row,column=self.id_column).data_type == 's':
                # remove hyphens
                #print row,self.id_column,val1
                val1 = val1.replace('-','')
                rowID = atoi(val1)
            else:
                rowID = int(val1)
            
            student = find_by_id(students,rowID)
            if student is not None:
                # Get Score
                val2 = self.s.cell(row=row,column=self.score_column).value
                score = 0
                if self.s.cell(row=row,column=self.score_column).data_type == 's':
                    score = atoi(val2)
                else:
                    score = int(val2)
                # Get Subject
                val3 = str(self.s.cell(row=row,column=self.course_column).value)
                subject = self.code_dict[val3[0:4]]
                # Get Month
                #print row, self.month_column, self.s.cell(row=row,column=self.month_column).value
                val4 = int(self.s.cell(row=row,column=self.month_column).value)
                if val4 == 1:   month = 1 # January Regents
                elif val4 == 2: month = 6 # June Regents
                elif val4 == 7: month = 8 # August Regents
                # Get Year
                val5 = int(str(self.s.cell(row=row,column=self.year_column).value)[0:2])
                year = val5 + 2000

                student.add_exam(Exam(examType='Regents',subject=subject,month=month,year=year,score=score))

class SATReader(Reader):
    def __init__(self,excelFile):
        super(SATReader, self).__init__(excelFile)

        self.header_row,self.id_column = self.find_id_cell()
        self.reading_column = self.find_column(['critical','reading'])
        self.math_column   = self.find_column(['math'])
        self.writing_column  = self.find_column(['writing'])
    
    def get_exams(self,students):
        for row in range(self.header_row+1,self.s.get_highest_row()+1):
            rowID = 0 
            val1 = self.s.cell(row=row,column=self.id_column).value
            if self.s.cell(row=row,column=self.id_column).data_type == 's':
                # remove hyphens
                val1 = val1.replace('-','')
                rowID = atoi(val1)
            else:
                rowID = int(val1)
            
            student = find_by_id(students,rowID)
            if student is not None:
                # FIXME: Hardcoded month and year until stored properly by DP
                month = 3
                year = 2012
                # Get Reading
                reading  = int(self.s.cell(row=row,column=self.reading_column).value)
                student.add_exam(Exam(examType='SAT',subject='Reading',month=month,year=year,score=reading))
                # Get Math
                math     = int(self.s.cell(row=row,column=self.math_column).value)
                student.add_exam(Exam(examType='SAT',subject='Mathematics',month=month,year=year,score=math))
                # Get Writing
                writing = int(self.s.cell(row=row,column=self.writing_column).value)
                student.add_exam(Exam(examType='SAT',subject='Writing',month=month,year=year,score=writing))


class ACTReader(Reader):
    def __init__(self,excelFile):
        super(ACTReader, self).__init__(excelFile)

        self.header_row,self.id_column = self.find_id_cell()
        self.english_column = self.find_column(['english'])
        self.reading_column = self.find_column(['reading'])
        self.math_column   = self.find_column(['math','mathematics'])
        self.writing_column  = self.find_column(['writing'])
        self.science_column  = self.find_column(['science'])
    
    def get_exams(self,students):
        for row in range(self.header_row+1,self.s.get_highest_row()+1):
            rowID = 0 
            val1 = self.s.cell(row=row,column=self.id_column).value
            if self.s.cell(row=row,column=self.id_column).data_type == 's':
                # remove hyphens
                val1 = val1.replace('-','')
                rowID = atoi(val1)
            else:
                rowID = int(val1)
            
            student = find_by_id(students,rowID)
            if student is not None:
                # FIXME: Hardcoded month and year unitl stored properly by DP
                month = 3
                year = 2012
                # Get English
                english  = int(self.s.cell(row=row,column=self.english_column).value)
                student.add_exam(Exam(examType='ACT',subject='English',month=month,year=year,score=english))
                # Get Reading
                reading  = int(self.s.cell(row=row,column=self.reading_column).value)
                student.add_exam(Exam(examType='ACT',subject='Reading',month=month,year=year,score=reading))
                # Get Math
                math     = int(self.s.cell(row=row,column=self.math_column).value)
                student.add_exam(Exam(examType='ACT',subject='Mathematics',month=month,year=year,score=math))
                # Get Science
                science = int(self.s.cell(row=row,column=self.science_column).value)
                student.add_exam(Exam(examType='ACT',subject='Science',month=month,year=year,score=science))
                if self.writing_column is not None:
                    # Get Writing
                    writing = int(self.s.cell(row=row,column=self.writing_column).value)
                    student.add_exam(Exam(examType='ACT',subject='Writing',month=month,year=year,score=writing))
    
class APReader(Reader):
    def __init__(self,excelFile):
        super(APReader, self).__init__(excelFile)

        self.header_row,self.id_column = self.find_id_cell()
        self.score_column  = self.find_column(['score'])
        self.code_column  = self.find_column(['code'])
        self.month_column  = self.find_column(['month'])
        self.year_column  = self.find_column(['year'])

        # Codes for AP Exams
        # Can be found here: http://tas.buffalo.edu/altexam/apchart.php
        self.code_dict = {
            'UH'   : 'US History',
            'EH'   : 'European History',
            'WH'   : 'World History',
            'ENGC' : 'English Lang. & Comp.',
            'ELC'  : 'English Lit. & Comp.',
            'BY'   : 'Biology',
            'CH'   : 'Chemistry',
            'PHB'  : 'Physics B',
            'PHCM' : 'Physics C: Mechanics',
            'PHCE' : 'Physics C: Elec. & Mag.',
            'MAB'  : 'Calculus AB',
            'MBC'  : 'Calculus BC',
            'STAT' : 'Statistics',
            'GPU'  : "US Gov't & Politics",
            'GPC'  : "Comp. Gov't & Politics"
        }
    
    def get_exams(self,students):
        for row in range(self.header_row+1,self.s.get_highest_row()+1):
            rowID = 0 
            val1 = self.s.cell(row=row,column=self.id_column).value
            if self.s.cell(row=row,column=self.id_column).data_type == 's':
                # remove hyphens
                val1 = val1.replace('-','')
                rowID = atoi(val1)
            else:
                rowID = int(val1)
            
            student = find_by_id(students,rowID)
            if student is not None:
                # Get Score
                score = int(self.s.cell(row=row,column=self.score_column).value)
                # Get Subject
                code = str(self.s.cell(row=row,column=self.code_column).value)
                subject = self.code_dict[code]
                # Get Month
                month = int(self.s.cell(row=row,column=self.month_column).value)
                # Get Year
                year = int(self.s.cell(row=row,column=self.year_column).value)

                student.add_exam(Exam(examType='AP',subject=subject,month=month,year=year,score=score))

def ReadMaster(docsFile, students,sections,target_grade = -1):
    gd_client = gdata.spreadsheet.service.SpreadsheetsService()
    email = raw_input('Enter DP username: (omit @democracyprep.org)  ')
    email += '@democracyprep.org'
    gd_client.email = email
    passwd = getpass.getpass(prompt='Enter Password: ')
    gd_client.password = passwd
    gd_client.source = 'DPGoogleSheet'
    gd_client.source = 'MasterSchedule'
    gd_client.ProgrammaticLogin()
    
    q = gdata.spreadsheet.service.DocumentQuery()
    
    q['title'] = docsFile 
    q['title-exact'] = 'true'
    feed = gd_client.GetSpreadsheetsFeed(query=q)
    spreadsheet_id = feed.entry[0].id.text.rsplit('/',1)[1]
    feed = gd_client.GetWorksheetsFeed(spreadsheet_id)
    worksheet_id = feed.entry[0].id.text.rsplit('/',1)[1]
    for worksheet in feed.entry:
        if 'MonThurSections' in str(worksheet.title):
            worksheet_id = worksheet.id.text.rsplit('/',1)[1]
            rows = gd_client.GetListFeed(spreadsheet_id, worksheet_id).entry
            
            for row in rows:
                #for key in row.custom:
                #    print " %s: %s" % (key, row.custom[key].text)
                title   = row.custom['course'].text
                grades   = row.custom['primarygrade'].text
                period   = row.custom['period'].text
                subject  = 'all'
                teacher = row.custom['teacher'].text
                if teacher == None: teacher = '-'
                room = row.custom['classroom'].text
                sections.append(Section(title=title,subject=subject,teacher=teacher,period=int(period),room=room))

    for worksheet in feed.entry:
        if 'FridaySection' in str(worksheet.title):
            worksheet_id = worksheet.id.text.rsplit('/',1)[1]
            rows = gd_client.GetListFeed(spreadsheet_id, worksheet_id).entry
            
            for row in rows:
                title   = row.custom['course'].text
                grades   = row.custom['primarygrade'].text
                period   = row.custom['period'].text
                subject  = 'all'
                teacher = row.custom['teacher'].text
                if teacher == None: teacher = '-'
                room = row.custom['classroom'].text
                sections.append(Section(title=title,subject=subject,teacher=teacher,period=int(period),room=room, friday=True))

    for worksheet in feed.entry:
        if 'Students' in str(worksheet.title):
            worksheet_id = worksheet.id.text.rsplit('/',1)[1]
            rows = gd_client.GetListFeed(spreadsheet_id, worksheet_id).entry
            
            for row in rows:
                #for key in row.custom:
                #    print " %s: %s" % (key, row.custom[key].text)
                grade   = row.custom['grade'].text
                if 'Leave' in grade: continue
                #if grade != target_grade and target_grade != -1: continue
                ID = row.custom['studentid'].text
                last =  row.custom['lastname'].text
                first =  row.custom['firstname'].text
                advisor =  row.custom['advisor'].text
                homeroom =  row.custom['homeroom'].text
                student = Student(stuID=ID,last=last,first=first,homeroom=homeroom,advisor=advisor)
                student.set_hs_grade(grade)
                students.append(student)

                for period in xrange(1,10):
                    sec_title = row.custom['period'+str(period)].text
                    if sec_title == None: continue
                    sec_found = False
                    for section in sections:
                        if sec_title == section.title and period == section.period and not section.is_friday:
                            section.students.append(student)
                            student.schedule[period-1] = section
                            sec_found = True

                    if not sec_found:
                        print 'WARNING: ',sec_title, 'not found in section catalog.'
                        tmp_section = Section(title=sec_title,subject='all',teacher='',period=int(period))
                        student.schedule[period-1] = tmp_section

                for period in xrange(1,7):
                    sec_title = row.custom['fridayperiod'+str(period)].text
                    sec_found = False
                    for section in sections:
                        if sec_title == section.title and period == section.period and section.is_friday:
                            section.students.append(student)
                            student.friday_schedule[period-1] = section
                            sec_found = True

                    if not sec_found:
                        print 'WARNING: ',sec_title, 'not found in friday section catalog.'
                        tmp_section = Section(title=sec_title,subject='all',teacher='',period=int(period),friday=True)
                        student.friday_schedule[period-1] = tmp_section

def ReadGoogleSheet(docsFile, sheetName, colNames):
    for i in xrange(0,len(colNames)):
        colNames[i] = (colNames[i].replace(' ','')).lower()
    gd_client = gdata.spreadsheet.service.SpreadsheetsService()
    email = raw_input('Enter DP username: (omit @democracyprep.org)  ')
    email += '@democracyprep.org'
    gd_client.email = email
    passwd = getpass.getpass(prompt='Enter Password: ')
    gd_client.password = passwd
    gd_client.source = 'DPGoogleSheet'
    gd_client.ProgrammaticLogin()
    
    q = gdata.spreadsheet.service.DocumentQuery()
    
    q['title'] = docsFile 
    q['title-exact'] = 'true'
    feed = gd_client.GetSpreadsheetsFeed(query=q)
    spreadsheet_id = feed.entry[0].id.text.rsplit('/',1)[1]
    feed = gd_client.GetWorksheetsFeed(spreadsheet_id)
    worksheet_id = feed.entry[0].id.text.rsplit('/',1)[1]
    data = []
    for worksheet in feed.entry:
        if sheetName in str(worksheet.title): #FIXME THis is stupid as this requires that all substrings be unique
            worksheet_id = worksheet.id.text.rsplit('/',1)[1]
            rows = gd_client.GetListFeed(spreadsheet_id, worksheet_id).entry
            i=0
            for row in rows:
                data.append([0] * len(colNames))
                idx = 0
                for key in colNames:
                    if key in row.custom.keys():
                        data[i][idx] = row.custom[key].text
                    elif i == 0:
                        print key,'not in',row.custom.keys()
                    idx += 1
                i += 1
    return data
